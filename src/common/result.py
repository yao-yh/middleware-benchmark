import time

import openpyxl
import logging
import os
logger = logging.getLogger('result')
base_path = 'output'
headers = [
    '测试名称',
    '测试开始时间',
    '总测试时长(s)',
    # 'iot-device-manager cpu_reservation',
    # 'iot-device-manager cpu_limit',
    # 'iot-device-manager mem_reservation',
    # 'iot-device-manager mem_limit',
    # 'iot-device-manager replicas',
    # 'emqx cpu_reservation',
    # 'emqx cpu_limit',
    # 'emqx mem_reservation',
    # 'emqx mem_limit',
    # 'emqx replicas',
]
class_map = {
    'emqx': 'EmqxClient',
    'kafka': 'KafkaClient',
    'kafka2': 'KafkaClient',
    # 'elasticsearch': 'ElasticsearchClient',
}

def init_result():
    """
    初始化解说文件
    """
    output_path = os.path.join(base_path, 'result.xlsx')
    if os.path.exists(output_path):
        return
    openpyxl.Workbook().save(output_path)
    xlsx = openpyxl.load_workbook(output_path)
    result = xlsx.active

    for key in class_map.keys():
        headers.append(f'{key} 消息个数')
        headers.append(f'{key} 消息个数停止更新时间')
        if key == 'mqtt':
            continue
        # headers.append(f'{key} 消息丢失(%)')
        # headers.append(f'{key} 最大时延(ms)')
        headers.append(f'{key} 平均时延（仅供参考）(ms)')

    for key in class_map.keys():
        headers.append(f'{key} 吞吐(条/s)')

    for col_num, value in enumerate(headers, start=1):
        result.cell(row=1, column=col_num, value=value)
    try:
        xlsx.save(output_path)
    except PermissionError:
        new_output_path = output_path + str(time.time())
        logger.error(f'文件权限不足，请关闭文件后重试。 已保存到 {new_output_path}')
        xlsx.save(new_output_path)

def save_to__result(row_result):
    output_path = os.path.join(base_path, 'result.xlsx')
    if not os.path.exists(output_path):
        openpyxl.Workbook().save(output_path)

    xlsx = openpyxl.load_workbook(output_path)
    result = xlsx.active

    result.append(row_result)

    xlsx.save('output/result.xlsx')
